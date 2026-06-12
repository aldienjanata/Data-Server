import openpyxl
import re
import uuid

EXCEL_FILE = r'Data Otb-Gtgo-Cisco Banyumas.xlsx'
OUTPUT_SQL = r'insert_banyumas.sql'

site_id = "uuid_generate_v4()"
site_code = "'BMS'"
site_name = "'Banyumas'"

def safe_str(val):
    if val is None: return "NULL"
    s = str(val).replace("'", "''").strip()
    if not s: return "NULL"
    return f"'{s}'"

def parse_otb_sheet(ws, device_name, is144=False):
    ports = []
    # OTB format parsing
    # row 3 onwards
    for row in ws.iter_rows(min_row=3, max_row=22, values_only=True):
        if not row[0] or not str(row[0]).endswith('.0'): continue
        
        row_idx = float(row[0])
        cores = row[1:]
        
        for c_idx, cell in enumerate(cores):
            if cell is None: continue
            
            c_str = str(cell).strip()
            if not c_str or c_str.lower() in ['core', 'c0re']: continue
            
            # Format: 'CORE X / Y' or 'CORE X'
            parts = c_str.split('/')
            core_part = parts[0].strip().upper().replace('CORE', '').replace('C0RE', '').strip()
            try:
                core_num = int(core_part)
            except:
                continue
            
            port_num = core_num
            if len(parts) > 1:
                try:
                    port_num = int(parts[1].strip())
                except:
                    pass
                    
            # In 144, port 1 is core 133 etc. We just map whatever port_num/core_num is in excel.
            # But wait, we just use the port_number directly.
            # Let's see if cell has a background color indicating it's filled?
            # openpyxl values_only=True loses color. We just need the structure.
            # Actually, the user just wants the structure inserted as 'empty'. They will fill it later.
            ports.append({
                'port_number': port_num,
                'core_number': core_num,
                'status': 'empty'
            })
            
    return ports

def parse_generic_sheet(ws, start_row, label_col_idx, num_col_idx):
    ports = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        label = row[label_col_idx]
        if not label: continue
        label = str(label).strip()
        if not label.lower().startswith('port'): continue
        
        try:
            port_num = int(re.search(r'\d+', label).group())
        except:
            continue
            
        ports.append({
            'port_number': port_num,
            'status': 'empty'
        })
    return ports

wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
devices = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    name = sheet_name.strip()
    upper_name = name.upper()
    
    device_type = 'OTHER'
    if 'OTB' in upper_name: device_type = 'OTB'
    elif 'CISCO' in upper_name: device_type = 'CISCO'
    elif 'HUAWEI' in upper_name: device_type = 'HUAWEI'
    elif 'GTGO' in upper_name: device_type = 'OLT'
    elif 'OLT' in upper_name: device_type = 'OLT'
    
    ports = []
    if device_type == 'OTB':
        ports = parse_otb_sheet(ws, name, '144' in upper_name)
    else:
        # Generic heuristic
        for start_row in range(1, 10):
            row_vals = [c.value for c in ws[start_row]]
            if any(v and 'PORT' in str(v).upper() for v in row_vals):
                label_idx = next(i for i, v in enumerate(row_vals) if v and 'PORT' in str(v).upper())
                ports = parse_generic_sheet(ws, start_row, label_idx, label_idx)
                break
                
    if not ports:
        # fallback to 48 ports
        ports = [{'port_number': i+1, 'status': 'empty'} for i in range(48)]
        
    devices.append({
        'name': name,
        'type_code': device_type,
        'ports': ports
    })

sql = []
sql.append("BEGIN;")
sql.append(f"-- SITE {site_name}")
sql.append(f"DO $$")
sql.append("DECLARE")
sql.append(f"  v_site_id uuid;")
sql.append("  v_type_id uuid;")
sql.append("  v_dev_id uuid;")
sql.append("BEGIN")
sql.append(f"  SELECT id INTO v_site_id FROM sites WHERE code = {site_code} LIMIT 1;")
sql.append(f"  IF v_site_id IS NULL THEN")
sql.append(f"    v_site_id := gen_random_uuid();")
sql.append(f"    INSERT INTO sites (id, name, code, location) VALUES (v_site_id, {site_name}, {site_code}, 'Jawa Tengah');")
sql.append(f"  END IF;")

for d in devices:
    d_name = safe_str(d['name'])
    t_code = safe_str(d['type_code'])
    
    sql.append(f"  -- DEVICE: {d['name']}")
    sql.append(f"  SELECT id INTO v_type_id FROM device_types WHERE name = {t_code} LIMIT 1;")
    sql.append(f"  v_dev_id := gen_random_uuid();")
    sql.append(f"  INSERT INTO devices (id, site_id, device_type_id, name, total_ports) VALUES (v_dev_id, v_site_id, v_type_id, {d_name}, {len(d['ports'])});")
    
    # insert ports
    if d['ports']:
        vals = []
        for p in d['ports']:
            cn = p.get('core_number')
            cn_str = str(cn) if cn else "NULL"
            vals.append(f"(v_dev_id, {p['port_number']}, {cn_str}, 'empty')")
            
        sql.append(f"  INSERT INTO port_connections (device_id, port_number, core_number, status) VALUES ")
        
        # chunking to avoid huge statements, though in PL/pgSQL we can just do one large insert if not too big
        chunk_size = 500
        for i in range(0, len(vals), chunk_size):
            chunk = vals[i:i+chunk_size]
            sql.append("    " + ",\n    ".join(chunk) + (";" if i+chunk_size >= len(vals) else ","))
            
sql.append("END $$;")
sql.append("COMMIT;")

with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))
print(f"SQL Generated: {OUTPUT_SQL}")
